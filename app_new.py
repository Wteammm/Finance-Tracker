from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, case, extract, text
from datetime import datetime
import os
import hashlib
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///finance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'finance_tracker_secret_key_2024'

# Password protection - SHA256 hash of "1234"
PASSWORD_HASH = '03ac674216f3e15c761ee1a5e255f067953623c8b388b4459e13f978d7c846f4'

db = SQLAlchemy(app)

# User Model for multi-user support
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self):
        return f'<User {self.username}>'

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    category = db.Column(db.String(50), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    account_id = db.Column(db.Integer, db.ForeignKey('balance_item.id'), nullable=True) # Link to Source Account

    def __repr__(self):
        return f'<Transaction {self.id} - {self.description}>'

class BalanceHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('balance_item.id'), nullable=False)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    old_value = db.Column(db.Float, nullable=False)
    new_value = db.Column(db.Float, nullable=False)
    adjustment = db.Column(db.Float, nullable=False)
    contra_account_id = db.Column(db.Integer, db.ForeignKey('balance_item.id'), nullable=True)
    contra_account_name = db.Column(db.String(100), nullable=True)
    description = db.Column(db.String(200), nullable=True)
    
    def __repr__(self):
        return f'<BalanceHistory {self.id} - Item {self.item_id}>'

class Investment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    type = db.Column(db.String(20), nullable=False) # Buy, Sell, Dividend
    symbol = db.Column(db.String(20), nullable=False)
    market = db.Column(db.String(20), nullable=False) # US, MY, Crypto, MMF
    quantity = db.Column(db.Float, nullable=False)
    price = db.Column(db.Float, nullable=False) # Price per unit in original currency
    fees = db.Column(db.Float, default=0.0) # Fees in original currency
    currency = db.Column(db.String(10), nullable=False) # USD, MYR
    exchange_rate = db.Column(db.Float, nullable=False, default=1.0) # To MYR (Used for MYR, ignored for USD if using avg rate)
    remark = db.Column(db.String(200), nullable=True) # New column
    brokerage = db.Column(db.String(100), nullable=True) # Brokerage account (e.g., Moomoo, Maybank Trade)

    def __repr__(self):
        return f'<Investment {self.symbol} - {self.type}>'

class ForexTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    amount_myr = db.Column(db.Float, nullable=False)
    rate = db.Column(db.Float, nullable=False) # MYR per USD
    amount_usd = db.Column(db.Float, nullable=False)

    def __repr__(self):
        return f'<Forex {self.amount_myr} MYR @ {self.rate}>'

class StockPrice(db.Model):
    symbol = db.Column(db.String(20), primary_key=True)
    price = db.Column(db.Float, nullable=False)
    last_updated = db.Column(db.DateTime, default=datetime.utcnow)

class PortfolioSnapshot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    total_invested_myr = db.Column(db.Float, nullable=False)
    total_value_myr = db.Column(db.Float, nullable=False)

class BalanceItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    classification = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    value = db.Column(db.Float, nullable=False)
    is_auto_linked = db.Column(db.Boolean, default=False)
    link_source = db.Column(db.String(50))
    liquidity_tier = db.Column(db.String(20), nullable=True)
    asset_type = db.Column(db.String(20), nullable=True)  # Cash, AR, Other
    obligation_type = db.Column(db.String(30), nullable=True, default='Standard')  # Standard, Held on Behalf, On Hold
    
    # Email reminder fields for AR
    contact_email = db.Column(db.String(200), nullable=True)
    contact_phone = db.Column(db.String(50), nullable=True)
    last_reminder_sent = db.Column(db.DateTime, nullable=True)
    reminder_count = db.Column(db.Integer, default=0)
    auto_reminder_enabled = db.Column(db.Boolean, default=False)  # Toggle auto reminders per contact


class Mortgage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    original_principal = db.Column(db.Float, nullable=False)
    term_years = db.Column(db.Integer, nullable=False)
    has_mrta = db.Column(db.Boolean, default=False)
    mrta_original_amount = db.Column(db.Float)
    mrta_rate = db.Column(db.Float)
    events = db.relationship('MortgageEvent', backref='mortgage', lazy=True)

class MortgageEvent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    mortgage_id = db.Column(db.Integer, db.ForeignKey('mortgage.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    type = db.Column(db.String(20), nullable=False) # 'PAYMENT', 'RATE_CHANGE'
    value = db.Column(db.Float, nullable=False)
    balance_after = db.Column(db.Float)

with app.app_context():
    db.create_all()
    # Migration for 'remark' column
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE investment ADD COLUMN remark VARCHAR(200)"))
            conn.commit()
    except Exception as e:
        pass # Column likely exists
    
    # Migration for 'brokerage' column
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE investment ADD COLUMN brokerage VARCHAR(100)"))
            conn.commit()
    except Exception as e:
        pass # Column likely exists
    
    # Migrations for AR email reminder columns
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE balance_item ADD COLUMN contact_email VARCHAR(200)"))
            conn.execute(text("ALTER TABLE balance_item ADD COLUMN contact_phone VARCHAR(50)"))
            conn.execute(text("ALTER TABLE balance_item ADD COLUMN last_reminder_sent DATETIME"))
            conn.execute(text("ALTER TABLE balance_item ADD COLUMN reminder_count INTEGER DEFAULT 0"))
            conn.execute(text("ALTER TABLE balance_item ADD COLUMN auto_reminder_enabled BOOLEAN DEFAULT 0"))
            conn.commit()
    except Exception as e:
        pass # Columns likely exist

def get_average_forex_rate():
    forex_txs = ForexTransaction.query.all()
    total_myr = sum(tx.amount_myr for tx in forex_txs)
    total_usd = sum(tx.amount_usd for tx in forex_txs)
    if total_usd == 0:
        return 1.0 # Default if no forex history
    return total_myr / total_usd

def get_portfolio_summary():
    investments = Investment.query.all()
    avg_forex_rate = get_average_forex_rate()
    total_invested_myr = 0
    
    for inv in investments:
        # Determine exchange rate to use
        if inv.currency == 'USD':
            rate = avg_forex_rate
        else:
            rate = inv.exchange_rate # Should be 1.0 for MYR
            
        # Calculate value in MYR
        # For Bonus/Split, cost is 0, so no addition to invested amount
        if inv.type in ['Buy']:
            val_myr = ((inv.price * inv.quantity) + inv.fees) * rate
            total_invested_myr += val_myr
        elif inv.type == 'Sell':
            # For Sell, we subtract the proceeds
            proceeds_myr = ((inv.price * inv.quantity) - inv.fees) * rate
            total_invested_myr -= proceeds_myr
        # Dividend, Bonus, Split do not affect "Invested Capital" directly in this simple view
        # Dividends are returns, not capital injection/removal (unless reinvested, which would be a Buy)
            
    return total_invested_myr


# ===== AUTHENTICATION =====

def login_required(f):
    """Decorator to protect routes - requires login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function



@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        # Validation
        if not username or not password:
            flash('Username and password are required.', 'danger')
            return redirect(url_for('register'))
        
        if password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return redirect(url_for('register'))
        
        # Check if username already exists
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Username already exists. Please choose another.', 'danger')
            return redirect(url_for('register'))
        
        # Create new user
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        new_user = User(username=username, password=hashed_password)
        db.session.add(new_user)
        db.session.commit()
        
        flash('Account created successfully! Please login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Check if user exists
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            # Successful login
            session['logged_in'] = True
            session['user_id'] = user.id
            session['username'] = user.username
            flash(f'Welcome back, {user.username}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password. Please try again.', 'danger')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ===== MAIN ROUTES =====

@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    if request.method == 'POST':
        date_str = request.form['date']
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date_obj = datetime.utcnow().date()
            
        category = request.form['category']
        description = request.form['description']
        try:
            amount = float(request.form['amount'])
        except ValueError:
            amount = 0.0
            
        account_id = request.form.get('account_id')
        if account_id and account_id.isdigit():
            account_id = int(account_id)
        else:
            account_id = None

        new_transaction = Transaction(date=date_obj, category=category, description=description, amount=amount, account_id=account_id)
        db.session.add(new_transaction)
        db.session.commit()
        return redirect(url_for('index'))

    # Filtering
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    
    query = Transaction.query
    if month and year:
        query = query.filter(extract('month', Transaction.date) == month, extract('year', Transaction.date) == year)
    elif year:
        query = query.filter(extract('year', Transaction.date) == year)
        
    transactions = query.order_by(Transaction.date.desc()).all()
    
    # Calculate Total Balance (Global)
    total_balance = sum(t.amount for t in transactions)
    
    # Get Accounts (BalanceItems) for Dropdown (Assets & Liabilities)
    accounts = BalanceItem.query.filter(BalanceItem.classification.in_(['Current Asset', 'Current Liability'])).all()
    
    # Calculate Live Balance for each Account
    for account in accounts:
        # Sum of linked transactions
        tx_sum = db.session.query(func.sum(Transaction.amount)).filter(Transaction.account_id == account.id).scalar() or 0.0
        # Live Balance = Initial Value + Transaction Sum
        account.current_balance = account.value + tx_sum
    
    # Get available years for filter
    years = db.session.query(extract('year', Transaction.date)).distinct().order_by(extract('year', Transaction.date).desc()).all()
    years = [int(y[0]) for y in years]
    
    current_month = month if month else datetime.now().month
    current_year = year if year else datetime.now().year

    return render_template('index.html', transactions=transactions, total_balance=total_balance, years=years, current_month=current_month, current_year=current_year, accounts=accounts)

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id):
    transaction = Transaction.query.get_or_404(id)
    if request.method == 'POST':
        date_str = request.form['date']
        try:
            transaction.date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
            
        transaction.category = request.form['category']
        transaction.description = request.form['description']
        try:
            transaction.amount = float(request.form['amount'])
        except ValueError:
            pass
            
        # Handle Account Link
        account_id = request.form.get('account_id')
        if account_id and account_id.isdigit():
            transaction.account_id = int(account_id)
        else:
            transaction.account_id = None

        db.session.commit()
        return redirect(url_for('index'))
        
    return render_template('edit.html', transaction=transaction)

@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    # print(f"DEBUG: Attempting to delete transaction {id}")
    transaction = Transaction.query.get_or_404(id)
    db.session.delete(transaction)
    db.session.commit()
    # print(f"DEBUG: Deleted transaction {id}")
    return redirect(url_for('index'))

@app.route('/mortgage/delete/<int:id>', methods=['POST'])
@login_required
def delete_mortgage(id):
    mortgage = Mortgage.query.get_or_404(id)
    # Delete associated events first
    MortgageEvent.query.filter_by(mortgage_id=id).delete()
    db.session.delete(mortgage)
    db.session.commit()
    return redirect(url_for('balance_sheet'))

@app.route('/chart')
@login_required
def chart():
    # Get available years for filter
    years = db.session.query(extract('year', Transaction.date)).distinct().order_by(extract('year', Transaction.date).desc()).all()
    years = [int(y[0]) for y in years]
    return render_template('chart.html', years=years)

@app.route('/chart-data')
@login_required
def chart_data():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)

    # Base query
    expenses_query = db.session.query(Transaction.category, func.sum(Transaction.amount)).filter(Transaction.amount < 0)
    income_query = db.session.query(Transaction.category, func.sum(Transaction.amount)).filter(Transaction.amount > 0)

    if month and year:
        expenses_query = expenses_query.filter(extract('month', Transaction.date) == month, extract('year', Transaction.date) == year)
        income_query = income_query.filter(extract('month', Transaction.date) == month, extract('year', Transaction.date) == year)
    elif year:
        expenses_query = expenses_query.filter(extract('year', Transaction.date) == year)
        income_query = income_query.filter(extract('year', Transaction.date) == year)

    expenses = expenses_query.group_by(Transaction.category).all()
    income = income_query.group_by(Transaction.category).all()

    return jsonify({
        'expenses': {category: abs(amount) for category, amount in expenses},
        'income': {category: amount for category, amount in income}
    })

@app.route('/portfolio')
@login_required
def portfolio():
    view_currency = request.args.get('currency', 'MYR')
    custom_rate_arg = request.args.get('custom_rate')
    
    avg_forex_rate = get_average_forex_rate()
    
    # Handle Custom Forex Rate (Current Market Rate)
    # Stored in StockPrice with symbol 'USDMYR'
    usd_myr_price = StockPrice.query.get('USDMYR')
    
    if custom_rate_arg:
        try:
            current_forex_rate = float(custom_rate_arg)
            if not usd_myr_price:
                usd_myr_price = StockPrice(symbol='USDMYR', price=current_forex_rate)
                db.session.add(usd_myr_price)
            else:
                usd_myr_price.price = current_forex_rate
            db.session.commit()
        except ValueError:
            current_forex_rate = usd_myr_price.price if usd_myr_price else avg_forex_rate
    else:
        current_forex_rate = usd_myr_price.price if usd_myr_price else avg_forex_rate
        
    investments = Investment.query.order_by(Investment.date.desc()).all()
    
    # Calculate Holdings - Group by (symbol, brokerage)
    holdings = {}
    first_buy_dates = {} # Track first buy date for annualized return

    for inv in investments:
        # Use tuple (symbol, brokerage) as key to separate holdings by brokerage
        key = (inv.symbol, inv.brokerage or '')
        
        if key not in holdings:
            holdings[key] = {
                'quantity': 0, 'cost_basis': 0, 'market': inv.market, 'currency': inv.currency,
                'total_dividends': 0, 'brokerage': inv.brokerage or ''
            }
        
        # Track first buy date per (symbol, brokerage)
        if inv.type == 'Buy':
            if key not in first_buy_dates or inv.date < first_buy_dates[key]:
                first_buy_dates[key] = inv.date

        if inv.type == 'Buy':
            holdings[key]['quantity'] += inv.quantity
            holdings[key]['cost_basis'] += (inv.price * inv.quantity) + inv.fees
        elif inv.type == 'Sell':
            if holdings[key]['quantity'] > 0:
                # Reduce cost basis proportionally
                avg_cost = holdings[key]['cost_basis'] / holdings[key]['quantity']
                cost_removed = avg_cost * inv.quantity
                holdings[key]['cost_basis'] -= cost_removed
                holdings[key]['quantity'] -= inv.quantity
        elif inv.type in ['Bonus', 'Split']:
            holdings[key]['quantity'] += inv.quantity
            # Cost basis does not change for Bonus/Split
        elif inv.type == 'Dividend':
            holdings[key]['total_dividends'] += inv.price # Price here stores the total dividend amount
            
    # Get current prices
    stock_prices = {sp.symbol: sp.price for sp in StockPrice.query.all()}

    # Prepare display data based on view_currency
    display_holdings = {}
    for key, data in holdings.items():
        if data['quantity'] > 0:
            symbol, brokerage = key  # Unpack tuple key
            display_data = data.copy()
            
            # Get Current Price (use symbol only, not brokerage)
            current_price = stock_prices.get(symbol, 0.0)
            display_data['current_price'] = current_price

            # Calculate Market Value (Native)
            market_value_native = data['quantity'] * current_price
            
            # Calculate Display Cost & Dividends & Market Value
            if view_currency == 'MYR':
                # COST uses AVG Rate (Historical)
                cost_rate = avg_forex_rate if data['currency'] == 'USD' else 1.0
                display_data['display_cost'] = data['cost_basis'] * cost_rate
                
                # VALUE uses CURRENT Rate (Market)
                value_rate = current_forex_rate if data['currency'] == 'USD' else 1.0
                display_data['display_market_value'] = market_value_native * value_rate
                
                # Dividends use Avg Rate (Realized) - keeping consistent with Cost for now
                div_rate = avg_forex_rate if data['currency'] == 'USD' else 1.0
                display_data['display_dividends'] = data['total_dividends'] * div_rate
                
                display_data['display_currency'] = 'MYR'
            elif view_currency == 'USD':
                rate = (1.0 / avg_forex_rate) if data['currency'] == 'MYR' and avg_forex_rate else 1.0
                if data['currency'] == 'USD': rate = 1.0 # Already USD
                
                display_data['display_cost'] = data['cost_basis'] * rate
                display_data['display_dividends'] = data['total_dividends'] * rate
                display_data['display_market_value'] = market_value_native * rate
                display_data['display_currency'] = 'USD'
            else: # Original
                display_data['display_cost'] = data['cost_basis']
                display_data['display_dividends'] = data['total_dividends']
                display_data['display_market_value'] = market_value_native
                display_data['display_currency'] = data['currency']
            
            # Calculate Yield on Cost
            if display_data['display_cost'] > 0:
                display_data['yield_on_cost'] = (display_data['display_dividends'] / display_data['display_cost']) * 100
            else:
                display_data['yield_on_cost'] = 0.0

            # Calculate P&L ($) and P&L (%)
            display_data['pnl_value'] = display_data['display_market_value'] - display_data['display_cost']
            if display_data['display_cost'] > 0:
                display_data['pnl_percent'] = (display_data['pnl_value'] / display_data['display_cost']) * 100
            else:
                display_data['pnl_percent'] = 0.0

            # Calculate Annualized Return (CAGR)
            # Formula: ((End Value + Dividends) / Start Value) ^ (365 / Days) - 1
            # Start Value = Cost Basis
            # End Value = Market Value
            if key in first_buy_dates and display_data['display_cost'] > 0:
                days_held = (datetime.utcnow().date() - first_buy_dates[key]).days
                if days_held > 0:
                    total_return_ratio = (display_data['display_market_value'] + display_data['display_dividends']) / display_data['display_cost']
                    try:
                        cagr = (total_return_ratio ** (365 / days_held)) - 1
                        display_data['annualized_return'] = cagr * 100
                    except:
                        display_data['annualized_return'] = 0.0
                else:
                     display_data['annualized_return'] = 0.0 # Held less than a day
            else:
                display_data['annualized_return'] = 0.0
                
            display_holdings[key] = display_data

    total_invested_myr = get_portfolio_summary()
    
    if view_currency == 'USD':
        total_invested_display = total_invested_myr / avg_forex_rate if avg_forex_rate else 0
        total_currency = 'USD'
    else:
        total_invested_display = total_invested_myr
        total_currency = 'MYR'

    return render_template('portfolio.html', 
                           investments=investments, 
                           holdings=display_holdings, 
                           total_invested=total_invested_display, 
                           total_currency=total_currency,
                           view_currency=view_currency,
                           avg_forex_rate=avg_forex_rate,
                           current_forex_rate=current_forex_rate)

@app.route('/portfolio/add', methods=['GET', 'POST'])
@login_required
def add_investment():
    if request.method == 'POST':
        date_str = request.form['date']
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date_obj = datetime.utcnow().date()
            
        type = request.form['type']
        symbol = request.form['symbol'].upper()
        market = request.form['market']
        quantity = float(request.form['quantity'])
        price = float(request.form['price'])
        fees = float(request.form['fees'])
        currency = request.form['currency']
        remark = request.form.get('remark', '')
        brokerage = request.form.get('brokerage', '')
        
        # Auto-set currency for Crypto to MYR
        if market == 'Crypto':
            currency = 'MYR'
        
        # Auto-set currency for MMF markets
        if market == 'MMF-RM':
            currency = 'MYR'
        elif market == 'MMF-USD':
            currency = 'USD'
        
        # Exchange rate is now 1.0 default, handled dynamically for USD
        exchange_rate = 1.0 

        new_inv = Investment(
            date=date_obj, type=type, symbol=symbol, market=market,
            quantity=quantity, price=price, fees=fees,
            currency=currency, exchange_rate=exchange_rate, remark=remark, brokerage=brokerage
        )
        db.session.add(new_inv)
        db.session.commit()
        return redirect(url_for('portfolio'))
    
    # Get all current holdings for Sell functionality
    holdings = db.session.query(
        Investment.symbol,
        Investment.market,
        Investment.currency,
        func.sum(case((Investment.type == 'Buy', Investment.quantity), else_=0)).label('bought'),
        func.sum(case((Investment.type == 'Sell', Investment.quantity), else_=0)).label('sold'),
        func.sum(case((Investment.type == 'Bonus', Investment.quantity), else_=0)).label('bonus'),
        func.sum(case((Investment.type == 'Split', Investment.quantity), else_=0)).label('split')
    ).group_by(Investment.symbol, Investment.market, Investment.currency).all()
    
    holdings_data = []
    for h in holdings:
        net_qty = h.bought - h.sold + h.bonus + h.split
        if net_qty > 0:
            holdings_data.append({
                'symbol': h.symbol,
                'market': h.market,
                'currency': h.currency,
                'quantity': net_qty
            })
    
    return render_template('add_investment.html', holdings=holdings_data)

@app.route('/portfolio/forex', methods=['GET', 'POST'])
@login_required
def forex():
    if request.method == 'POST':
        date_str = request.form['date']
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date_obj = datetime.utcnow().date()
            
        amount_myr = float(request.form['amount_myr'])
        rate = float(request.form['rate'])
        amount_usd = amount_myr / rate
        
        new_forex = ForexTransaction(date=date_obj, amount_myr=amount_myr, rate=rate, amount_usd=amount_usd)
        db.session.add(new_forex)
        db.session.commit()
        return redirect(url_for('forex'))
        
    forex_txs = ForexTransaction.query.order_by(ForexTransaction.date.desc()).all()
    avg_rate = get_average_forex_rate()
    
    return render_template('forex.html', forex_txs=forex_txs, avg_rate=avg_rate)

@app.route('/portfolio/prices', methods=['GET', 'POST'])
@login_required
def update_prices():
    # Get all unique symbols
    symbols = db.session.query(Investment.symbol).distinct().all()
    symbols = [s[0] for s in symbols]
    
    if request.method == 'POST':
        if request.form.get('single_update'):
            # Handle single price update from modal
            symbol = request.form.get('symbol')
            price_str = request.form.get('price')
            if symbol and price_str:
                try:
                    price = float(price_str)
                    stock_price = StockPrice.query.get(symbol)
                    if not stock_price:
                        stock_price = StockPrice(symbol=symbol, price=price)
                        db.session.add(stock_price)
                    else:
                        stock_price.price = price
                        stock_price.last_updated = datetime.utcnow()
                except ValueError:
                    pass
        else:
            # Handle bulk update
            for symbol in symbols:
                price_str = request.form.get(f'price_{symbol}')
                if price_str:
                    try:
                        price = float(price_str)
                        stock_price = StockPrice.query.get(symbol)
                        if not stock_price:
                            stock_price = StockPrice(symbol=symbol, price=price)
                            db.session.add(stock_price)
                        else:
                            stock_price.price = price
                            stock_price.last_updated = datetime.utcnow()
                    except ValueError:
                        pass
        db.session.commit()
        return redirect(url_for('portfolio'))
        
    # Get current prices
    current_prices = {sp.symbol: sp.price for sp in StockPrice.query.all()}
    return render_template('update_prices.html', symbols=symbols, current_prices=current_prices)

@app.route('/edit_investment/<int:id>', methods=['POST'])
@login_required
def edit_investment(id):
    inv = Investment.query.get_or_404(id)
    remark = request.form.get('remark', '')
    brokerage = request.form.get('brokerage', '')
    inv.remark = remark
    inv.brokerage = brokerage
    db.session.commit()
    return redirect(url_for('portfolio'))

@app.route('/delete_investment/<int:id>', methods=['POST'])
@login_required
def delete_investment(id):
    inv = Investment.query.get_or_404(id)
    db.session.delete(inv)
    db.session.commit()
    return redirect(url_for('portfolio'))

    return redirect(url_for('portfolio'))

# --- Balance Sheet Logic ---

def calculate_mortgage_balance(mortgage):
    """
    Calculates the current balance of a mortgage by replaying events.
    """
    current_balance = mortgage.original_principal
    # Events should be ordered by date
    events = sorted(mortgage.events, key=lambda x: x.date)
    
    # In a full implementation, we'd calculate daily interest.
    # For now, we assume simple principal reduction from payments.
    # The user manual entry or detailed scheduler will handle interest adjustments.
    # This function returns the "Principal Outstanding".
    
    for event in events:
        if event.type == 'PAYMENT':
             # Assumption: event.value is the principal portion paid? 
             # Or total payment? 
             # For Balance Sheet Summary, we need Principal Balance.
             # Let's assume for now the user records "Principal Paid" or we track balance_after.
             if event.balance_after is not None:
                 current_balance = event.balance_after
             else:
                 # Fallback if balance_after not set (e.g. simple payment subtraction)
                 # This is a simplification; verify with user preference if they want interest calc here.
                 pass 
                 
    # If no events, balance is original principal.
    # If events exist, we take the last known balance.
    if events and events[-1].balance_after:
        return events[-1].balance_after
        
    return current_balance

def add_months(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, [31,
        29 if year % 4 == 0 and not year % 100 == 0 or year % 400 == 0 else 28,
        31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return datetime(year, month, day).date()

def generate_mortgage_schedule(mortgage):
    schedule = []
    
    # 1. Timeline & Events
    start_date = mortgage.start_date
    end_date = add_months(start_date, mortgage.term_years * 12)
    today = datetime.utcnow().date()
    
    events = sorted(mortgage.events, key=lambda x: x.date)
    
    # Track Simulation State
    curr_balance = mortgage.original_principal
    curr_rate = 0.0
    
    # Initial Rate
    init_rate_event = next((e for e in events if e.type == 'RATE_CHANGE'), None)
    if init_rate_event:
        curr_rate = init_rate_event.value
        
    # MRTA State - Fixed Schedule based on inception
    mrta_balance = mortgage.mrta_original_amount if mortgage.has_mrta else 0
    mrta_rate = mortgage.mrta_rate if mortgage.has_mrta else 0
    # Calculate fixed monthly PMT for MRTA schedule
    mrta_pmt = 0
    if mortgage.has_mrta and mrta_rate > 0:
        r = mrta_rate / 100 / 12
        n = mortgage.term_years * 12
        if r > 0:
            mrta_pmt = (mrta_balance * r * (1 + r)**n) / ((1 + r)**n - 1)
        else:
            mrta_pmt = mrta_balance / n

    # --- HISTORICAL REPLAY ---
    # We iterate month by month from start_date until today
    
    # Naive assumption: Payments happen on the same day of month as start_date
    # Better approach: Just list all months, and see if there was a PAYMENT event in that month?
    # OR: Just show the official "Due Dates" and map actual payments to them.
    # To simplify: We just generate a monthly schedule.
    
    total_months = mortgage.term_years * 12
    
    last_event_idx = 0
    
    for i in range(total_months + 1): # +1 to show opening
        date = add_months(start_date, i)
        
        # Process events up to this date
        while last_event_idx < len(events) and events[last_event_idx].date <= date:
            evt = events[last_event_idx]
            if evt.type == 'RATE_CHANGE':
                curr_rate = evt.value
            elif evt.type == 'PAYMENT':
                # Payment reduces balance immediately
                # Logic: We display the payment in the schedule row corresponding to the month
                pass 
            last_event_idx += 1

        # Calculate MRTA Balance for this month (Standard Amortization)
        # B_k = B_0 * ... formula or iterative
        # Iterative is safer
        if i > 0 and mortgage.has_mrta:
             interest = mrta_balance * (mrta_rate / 100 / 12)
             principal_part = mrta_pmt - interest
             mrta_balance -= principal_part
             if mrta_balance < 0: mrta_balance = 0
        
        row = {
            'no': i,
            'date': date,
            'rate': curr_rate,
            'mrta_coverage': mrta_balance if mortgage.has_mrta else 0,
            'net_exposure': max(0, curr_balance - (mrta_balance if mortgage.has_mrta else 0))
        }

        # Handling Actuals vs Future
        if date <= today:
            row['type'] = 'History'
            row['balance'] = curr_balance # Should we try to find the EXACT balance at this date?
            # Ideally we recalculate balance month-by-month.
            # But we have `calculate_mortgage_balance` which gives valid snapshot.
            # Simplified: For history, we just show "Current Balance" as it stands. 
            # (Refining this requires a daily interest calculator which is too complex for this step).
            # Current simplified model: Events dictate balance.
            
            # Find payment in this month
            payment_evt = next((e for e in mortgage.events if e.type=='PAYMENT' and e.date.year==date.year and e.date.month==date.month), None)
            if payment_evt:
                row['payment'] = payment_evt.value
                row['balance'] = payment_evt.balance_after
                curr_balance = payment_evt.balance_after # Update runner
            else:
                row['payment'] = 0
                row['balance'] = curr_balance
                
            row['interest_paid'] = 0 # Todo: calculate interest component
            row['principal_paid'] = 0
            
        else:
            row['type'] = 'Projected'
            # Calculate PMT for remaining balance
            # Remaining n
            remaining_months = total_months - i + 1
            if remaining_months <= 0: break
            
            r = curr_rate / 100 / 12
            if r > 0 and curr_balance > 0:
                pmt = (curr_balance * r * (1 + r)**remaining_months) / ((1 + r)**remaining_months - 1)
            else:
                pmt = curr_balance / remaining_months if remaining_months > 0 else 0
            
            interest = curr_balance * r
            principal = pmt - interest
            curr_balance -= principal
            if curr_balance < 0: curr_balance = 0
            
            row['payment'] = pmt
            row['interest_paid'] = interest
            row['principal_paid'] = principal
            row['balance'] = curr_balance
            row['net_exposure'] = max(0, curr_balance - row['mrta_coverage'])
            
        schedule.append(row)
        if curr_balance <= 0 and i > 0: break # Paid off
        
    return schedule

@app.route('/mortgage/<int:id>', methods=['GET'])
@login_required
def mortgage_detail(id):
    mortgage = Mortgage.query.get_or_404(id)
    
    # Generate Schedule
    schedule = generate_mortgage_schedule(mortgage)
    
    # Get current snapshot
    current_balance = schedule[-1]['balance'] # Or calculate strictly
    # Better: use helper
    current_balance = calculate_mortgage_balance(mortgage)
    
    # Current Rate
    rate_events = [e for e in mortgage.events if e.type == 'RATE_CHANGE']
    rate_events.sort(key=lambda x: x.date)
    current_rate = rate_events[-1].value if rate_events else 0.0
    
    # MRTA info
    current_mrta = 0
    if mortgage.has_mrta:
        today = datetime.utcnow().date()
        # Find schedule row closest to today
        # simplified:
        for row in schedule:
            if row['date'] > today:
                current_mrta = row['mrta_coverage']
                break
        else:
            current_mrta = schedule[-1]['mrta_coverage'] if schedule else 0

    return render_template('mortgage_detail.html', 
                           mortgage=mortgage, 
                           current_balance=current_balance,
                           current_rate=current_rate,
                           current_mrta=current_mrta,
                           net_exposure=max(0, current_balance - current_mrta),
                           schedule=schedule)

@app.route('/mortgage/<int:id>/update_rate', methods=['POST'])
@login_required
def update_mortgage_rate(id):
    mortgage = Mortgage.query.get_or_404(id)
    new_rate = float(request.form['new_rate'])
    date_str = request.form['date']
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # We need to know the balance at this date to ensure continuity?
    # Not strictly, but helps. 
    # For now, just record the event. The scheduler handles logic.
    
    evt = MortgageEvent(mortgage_id=id, date=date_obj, type='RATE_CHANGE', value=new_rate)
    db.session.add(evt)
    db.session.commit()
    return redirect(url_for('mortgage_detail', id=id))

@app.route('/mortgage/<int:id>/payment', methods=['POST'])
@login_required
def add_mortgage_payment(id):
    mortgage = Mortgage.query.get_or_404(id)
    amount = float(request.form['amount'])
    date_str = request.form['date']
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # Calculate balance after this payment
    # This requires replaying all events up to this date
    # Simplification: Fetch current balance, subtract amount.
    current_bal = calculate_mortgage_balance(mortgage)
    new_bal = current_bal - amount # Assuming principal payment mainly
    # If standard monthly payment (Principal + Interest), we need to split.
    # For MVP: User enters "Principal Paid" or we handle total payment logic later.
    # Let's assume user enters TOTAL payment and we deduce? No, too complex.
    # Let's assume user enters PRINCIPAL reduction (Safe).
    # OR: just record balance_after manually? No, user doesn't know.
    
    # MVP approach: Just subtract from current balance. Refine later.
    
    evt = MortgageEvent(
        mortgage_id=id, date=date_obj, type='PAYMENT', 
        value=amount, balance_after=new_bal
    )
    db.session.add(evt)
    db.session.commit()
    return redirect(url_for('mortgage_detail', id=id))

@app.route('/balance_sheet', methods=['GET', 'POST'])
@login_required
def balance_sheet():
    if request.method == 'POST':
        # Add Manual Item
        classification = request.form['classification']
        name = request.form['name']
        asset_type = request.form.get('asset_type')
        liquidity_tier = request.form.get('liquidity_tier')
        
        # Email reminder fields for AR
        contact_email = request.form.get('contact_email')
        contact_phone = request.form.get('contact_phone')
        auto_reminder_enabled = request.form.get('auto_reminder_enabled') == 'on'
        
        try:
            value = float(request.form['value'])
        except ValueError:
            value = 0.0
            
        new_item = BalanceItem(
            classification=classification, 
            name=name, 
            value=value, 
            is_auto_linked=False,
            asset_type=asset_type,
            liquidity_tier=liquidity_tier,
            contact_email=contact_email if asset_type == 'AR' else None,
            contact_phone=contact_phone if asset_type == 'AR' else None,
            auto_reminder_enabled=auto_reminder_enabled if asset_type == 'AR' else False
        )
        db.session.add(new_item)
        db.session.commit()
        return redirect(url_for('balance_sheet'))

    # 1. Get Auto-Linked Cash (Current Asset) - Unallocated Only
    # Only sum transactions that are NOT linked to a specific account
    transactions = Transaction.query.all()
    total_cash = sum(t.amount for t in transactions if t.account_id is None)
    
    # 2. Get Auto-Linked Investments (Non-Current Asset)
    # Using existing logic logic from portfolio_overview to get Market Value in MYR
    investments = Investment.query.all()
    avg_forex_rate = get_average_forex_rate()
    usd_myr_price = StockPrice.query.get('USDMYR')
    current_forex_rate = usd_myr_price.price if usd_myr_price else avg_forex_rate
    stock_prices = {sp.symbol: sp.price for sp in StockPrice.query.all()}
    
    total_portfolio_value_myr = 0
    positions = {}
    
    # Quick calc loop (simplified from portfolio_overview)
    for inv in investments:
        if inv.symbol not in positions:
            positions[inv.symbol] = {'quantity': 0, 'market': inv.market, 'currency': inv.currency}
        
        if inv.type == 'Buy': positions[inv.symbol]['quantity'] += inv.quantity
        elif inv.type == 'Sell': positions[inv.symbol]['quantity'] -= inv.quantity
        elif inv.type in ['Bonus', 'Split']: positions[inv.symbol]['quantity'] += inv.quantity
            
    # Group positions by market
    market_totals = {'US': 0.0, 'MY': 0.0, 'Crypto': 0.0, 'MMF': 0.0}
    
    for symbol, pos in positions.items():
        if pos['quantity'] > 0:
            price = stock_prices.get(symbol, 0.0)
            val = pos['quantity'] * price
            # Convert to MYR using Current Rate for Value
            if pos['currency'] == 'USD':
                val *= current_forex_rate
            total_portfolio_value_myr += val
            
            # Add to market total
            market = pos.get('market', 'Other')
            if market in market_totals:
                market_totals[market] += val

    # 3. Get Mortgages (Non-Current Liability)
    mortgages = Mortgage.query.all()
    total_mortgage_balance = 0
    mortgage_data = []
    for m in mortgages:
        bal = calculate_mortgage_balance(m)
        total_mortgage_balance += bal
        mortgage_data.append({'name': m.name, 'balance': bal, 'id': m.id})

    # 4. Get Manual Items and Calculate Dynamic Balances
    manual_items = BalanceItem.query.all()
    
    # Pre-fetch transaction sums per account
    account_sums = db.session.query(
        Transaction.account_id, func.sum(Transaction.amount)
    ).filter(Transaction.account_id.isnot(None)).group_by(Transaction.account_id).all()
    account_sum_map = {acc_id: total for acc_id, total in account_sums}
    
    
    # Aggregation
    assets = {
        'current': [],  # Flat list for all current assets
        'non_current': {
            'investment_portfolio': {
                'total': total_portfolio_value_myr,
                'breakdown': [
                    {'name': 'US Stocks', 'value': market_totals['US'], 'is_auto': True},
                    {'name': 'MY Stocks', 'value': market_totals['MY'], 'is_auto': True},
                    {'name': 'Crypto', 'value': market_totals['Crypto'], 'is_auto': True},
                    {'name': 'MMF', 'value': market_totals['MMF'], 'is_auto': True}
                ]
            },
            'manual_items': []
        }
    }
    liabilities = {
        'current': [],
        'non_current': []
        # Mortgage added below
    }
    
    # Add Mortgages to Non-Current Liabilities
    for m in mortgage_data:
        liabilities['non_current'].append({'name': f"{m['name']} (Mortgage)", 'value': m['balance'], 'is_auto': True, 'link': f"/mortgage/{m['id']}"})
        
    # Categorize Manual Items
    for item in manual_items:
        # Calculate dynamic value if linked transactions exist
        current_value = item.value
        transaction_sum = account_sum_map.get(item.id, 0.0)
        final_value = current_value + transaction_sum
        
        entry = {'name': item.name, 'value': final_value, 'is_auto': False, 'id': item.id, 'liquidity_tier': item.liquidity_tier, 'asset_type': item.asset_type}
        
        if item.classification == 'Current Asset':
            # Add all current assets to flat list
            assets['current'].append(entry)
                
        elif item.classification == 'Non-Current Asset':
            assets['non_current']['manual_items'].append(entry)
        elif item.classification == 'Current Liability':
            liabilities['current'].append(entry)
        elif item.classification == 'Non-Current Liability':
            liabilities['non_current'].append(entry)
            
    # Add auto-generated unallocated cash to current assets
    if total_cash > 0:
        assets['current'].append({'name': 'Unallocated Cash', 'value': total_cash, 'is_auto': True, 'asset_type': 'Cash', 'liquidity_tier': None})
    
    # Group current assets by type with subtotals
    from collections import defaultdict
    grouped_current = defaultdict(lambda: {'item_list': [], 'total': 0})
    
    for item in assets['current']:
        type_key = item.get('asset_type') or 'Other'
        grouped_current[type_key]['item_list'].append(item)
        grouped_current[type_key]['total'] += item['value']
    
    # Replace flat list with grouped dictionary
    assets['current_grouped'] = dict(grouped_current)
    
    # Calculate Totals
    # Sum up all current assets from flat list
    total_current_assets = sum(item['value'] for item in assets['current'])
        
    # Total = Investment Portfolio + Manual Items
    total_non_current_assets = assets['non_current']['investment_portfolio']['total'] + sum(item['value'] for item in assets['non_current']['manual_items'])
    total_current_liabilities = sum(item['value'] for item in liabilities['current'])
    total_non_current_liabilities = sum(item['value'] for item in liabilities['non_current'])
    
    total_assets = total_current_assets + total_non_current_assets
    
    total_current_liabilities = sum(x['value'] for x in liabilities['current'])
    total_non_current_liabilities = sum(x['value'] for x in liabilities['non_current'])
    total_liabilities = total_current_liabilities + total_non_current_liabilities
    
    net_worth = total_assets - total_liabilities
    
    return render_template('balance_sheet.html',
                           assets=assets,
                           liabilities=liabilities,
                           total_current_assets=total_current_assets,
                           total_non_current_assets=total_non_current_assets,
                           total_assets=total_assets,
                           total_current_liabilities=total_current_liabilities,
                           total_non_current_liabilities=total_non_current_liabilities,
                           total_liabilities=total_liabilities,
                           net_worth=net_worth)

@app.route('/cash_flow')
@login_required
def cash_flow():
    # Reuse balance sheet data - get manual and auto items
    transactions = Transaction.query.all()
    total_cash_auto = sum(t.amount for t in transactions if t.account_id is None)
    
    # Get manual items with transaction sums
    manual_items = BalanceItem.query.all()
    account_sums = db.session.query(
        Transaction.account_id, func.sum(Transaction.amount)
    ).filter(Transaction.account_id.isnot(None)).group_by(Transaction.account_id).all()
    account_sum_map = {acc_id: total for acc_id, total in account_sums}
    
    # Build current assets list
    current_assets = []
    current_liabilities = []
    
    for item in manual_items:
        current_value = item.value
        transaction_sum = account_sum_map.get(item.id, 0.0)
        final_value = current_value + transaction_sum
        
        entry = {
            'name': item.name,
            'value': final_value,
            'is_auto': False,
            'id': item.id,
            'liquidity_tier': item.liquidity_tier,
            'asset_type': item.asset_type,
            'obligation_type': item.obligation_type or 'Standard'
        }
        
        if item.classification == 'Current Asset':
            current_assets.append(entry)
        elif item.classification == 'Current Liability':
            current_liabilities.append(entry)
    
    # Add auto unallocated cash
    if total_cash_auto > 0:
        current_assets.append({
            'name': 'Unallocated Cash',
            'value': total_cash_auto,
            'is_auto': True,
            'asset_type': 'Cash',
            'liquidity_tier': None
        })
    
    # Group current assets by type
    from collections import defaultdict
    grouped_current = defaultdict(lambda: {'item_list': [], 'total': 0})
    
    for item in current_assets:
        type_key = item.get('asset_type') or 'Other'
        grouped_current[type_key]['item_list'].append(item)
        grouped_current[type_key]['total'] += item['value']
    
    # Group liabilities by obligation type
    grouped_liabilities = defaultdict(lambda: {'item_list': [], 'total': 0})
    
    for item in current_liabilities:
        obligation_key = item.get('obligation_type') or 'Standard'
        grouped_liabilities[obligation_key]['item_list'].append(item)
        grouped_liabilities[obligation_key]['total'] += item['value']
    
    # Calculate cash flow metrics
    total_cash = grouped_current.get('Cash', {}).get('total', 0)
    total_ar = grouped_current.get('AR', {}).get('total', 0)
    total_obligations = sum(item['value'] for item in current_liabilities)
    
    # Calculate obligation type breakdowns
    held_on_behalf = grouped_liabilities.get('Held on Behalf', {}).get('total', 0)
    on_hold = grouped_liabilities.get('On Hold', {}).get('total', 0)
    standard_obligations = grouped_liabilities.get('Standard', {}).get('total', 0)
    
    available_cash = total_cash - total_obligations
    expected_cash = available_cash + total_ar
    
    return render_template('cash_flow.html',
                           total_cash=total_cash,
                           total_ar=total_ar,
                           total_obligations=total_obligations,
                           held_on_behalf=held_on_behalf,
                           on_hold=on_hold,
                           standard_obligations=standard_obligations,
                           available_cash=available_cash,
                           expected_cash=expected_cash,
                           cash_items=grouped_current.get('Cash', {}).get('item_list', []),
                           ar_items=grouped_current.get('AR', {}).get('item_list', []),
                           liability_items=current_liabilities,
                           grouped_liabilities=dict(grouped_liabilities))


@app.route('/cash_flow/export')
@login_required
def cash_flow_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file
    import datetime
    
    # Reuse cash_flow logic to get data
    transactions = Transaction.query.all()
    total_cash_auto = sum(t.amount for t in transactions if t.account_id is None)
    
    manual_items = BalanceItem.query.all()
    account_sums = db.session.query(
        Transaction.account_id, func.sum(Transaction.amount)
    ).filter(Transaction.account_id.isnot(None)).group_by(Transaction.account_id).all()
    account_sum_map = {acc_id: total for acc_id, total in account_sums}
    
    current_assets = []
    current_liabilities = []
    
    for item in manual_items:
        current_value = item.value
        transaction_sum = account_sum_map.get(item.id, 0.0)
        final_value = current_value + transaction_sum
        
        entry = {
            'name': item.name,
            'value': final_value,
            'liquidity_tier': item.liquidity_tier,
            'asset_type': item.asset_type,
            'obligation_type': item.obligation_type or 'Standard'
        }
        
        if item.classification == 'Current Asset':
            current_assets.append(entry)
        elif item.classification == 'Current Liability':
            current_liabilities.append(entry)
    
    if total_cash_auto > 0:
        current_assets.append({
            'name': 'Unallocated Cash',
            'value': total_cash_auto,
            'asset_type': 'Cash',
            'liquidity_tier': None
        })
    
    from collections import defaultdict
    grouped_current = defaultdict(lambda: {'item_list': [], 'total': 0})
    
    for item in current_assets:
        type_key = item.get('asset_type') or 'Other'
        grouped_current[type_key]['item_list'].append(item)
        grouped_current[type_key]['total'] += item['value']
    
    grouped_liabilities = defaultdict(lambda: {'item_list': [], 'total': 0})
    
    for item in current_liabilities:
        obligation_key = item.get('obligation_type') or 'Standard'
        grouped_liabilities[obligation_key]['item_list'].append(item)
        grouped_liabilities[obligation_key]['total'] += item['value']
    
    total_cash = grouped_current.get('Cash', {}).get('total', 0)
    total_ar = grouped_current.get('AR', {}).get('total', 0)
    total_obligations = sum(item['value'] for item in current_liabilities)
    held_on_behalf = grouped_liabilities.get('Held on Behalf', {}).get('total', 0)
    on_hold = grouped_liabilities.get('On Hold', {}).get('total', 0)
    standard_obligations = grouped_liabilities.get('Standard', {}).get('total', 0)
    available_cash = total_cash - total_obligations
    expected_cash = available_cash + total_ar
    
    # Create Excel workbook
    wb = Workbook()
    
    # Gold and black theme colors
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    black_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    gold_font = Font(color="D4AF37", bold=True)
    
    # Summary Sheet
    ws1 = wb.active
    ws1.title = "Cash Flow Summary"
    
    ws1['A1'] = "Cash Flow Overview"
    ws1['A1'].font = Font(size=16, bold=True, color="D4AF37")
    ws1['A2'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    ws1['A4'] = "Metric"
    ws1['B4'] = "Amount (RM)"
    ws1['A4'].fill = gold_fill
    ws1['B4'].fill = gold_fill
    ws1['A4'].font = white_font
    ws1['B4'].font = white_font
    
    metrics = [
        ("Available Cash", available_cash),
        ("", ""),
        ("Total Cash", total_cash),
        ("Current Obligations", -total_obligations),
        ("Accounts Receivable", total_ar),
        ("Expected Cash", expected_cash),
        ("", ""),
        ("Obligation Breakdown:", ""),
        ("  - Held on Behalf", held_on_behalf),
        ("  - On Hold", on_hold),
        ("  - Standard Debt", standard_obligations)
    ]
    
    for idx, (metric, value) in enumerate(metrics, start=5):
        ws1[f'A{idx}'] = metric
        if value != "":
            ws1[f'B{idx}'] = value
            ws1[f'B{idx}'].number_format = '#,##0.00'
    
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    # Cash Breakdown Sheet
    ws2 = wb.create_sheet("Cash Breakdown")
    ws2['A1'] = "Cash Breakdown"
    ws2['A1'].font = gold_font
    
    ws2['A3'] = "Item"
    ws2['B3'] = "Liquidity Tier"
    ws2['C3'] = "Amount (RM)"
    for cell in [ws2['A3'], ws2['B3'], ws2['C3']]:
        cell.fill = gold_fill
        cell.font = white_font
    
    row = 4
    for item in grouped_current.get('Cash', {}).get('item_list', []):
        ws2[f'A{row}'] = item['name']
        ws2[f'B{row}'] = item.get('liquidity_tier', 'N/A')
        ws2[f'C{row}'] = item['value']
        ws2[f'C{row}'].number_format = '#,##0.00'
        row += 1
    
    ws2[f'A{row}'] = "TOTAL"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'C{row}'] = total_cash
    ws2[f'C{row}'].number_format = '#,##0.00'
    ws2[f'C{row}'].font = Font(bold=True, color="D4AF37")
    
    # Obligations Breakdown Sheet
    ws3 = wb.create_sheet("Obligations Breakdown")
    ws3['A1'] = "Current Obligations"
    ws3['A1'].font = gold_font
    
    ws3['A3'] = "Type"
    ws3['B3'] = "Item"
    ws3['C3'] = "Amount (RM)"
    for cell in [ws3['A3'], ws3['B3'], ws3['C3']]:
        cell.fill = gold_fill
        cell.font = white_font
    
    row = 4
    for obligation_type, group_data in grouped_liabilities.items():
        ws3[f'A{row}'] = obligation_type
        ws3[f'A{row}'].font = Font(bold=True)
        ws3[f'C{row}'] = group_data['total']
        ws3[f'C{row}'].number_format = '#,##0.00'
        ws3[f'C{row}'].font = Font(bold=True)
        row += 1
        
        for item in group_data['item_list']:
            ws3[f'B{row}'] = item['name']
            ws3[f'C{row}'] = item['value']
            ws3[f'C{row}'].number_format = '#,##0.00'
            row += 1
        row += 1
    
    # AR Breakdown Sheet
    ws4 = wb.create_sheet("Accounts Receivable")
    ws4['A1'] = "Accounts Receivable"
    ws4['A1'].font = gold_font
    
    ws4['A3'] = "Item"
    ws4['B3'] = "Liquidity Tier"
    ws4['C3'] = "Amount (RM)"
    for cell in [ws4['A3'], ws4['B3'], ws4['C3']]:
        cell.fill = gold_fill
        cell.font = white_font
    
    row = 4
    for item in grouped_current.get('AR', {}).get('item_list', []):
        ws4[f'A{row}'] = item['name']
        ws4[f'B{row}'] = item.get('liquidity_tier', 'N/A')
        ws4[f'C{row}'] = item['value']
        ws4[f'C{row}'].number_format = '#,##0.00'
        row += 1
    
    ws4[f'A{row}'] = "TOTAL"
    ws4[f'A{row}'].font = Font(bold=True)
    ws4[f'C{row}'] = total_ar
    ws4[f'C{row}'].number_format = '#,##0.00'
    ws4[f'C{row}'].font = Font(bold=True, color="D4AF37")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cash_flow_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/export_transactions')
@login_required
def export_transactions():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import send_file
    import datetime
    
    # Get all transactions
    transactions = Transaction.query.order_by(Transaction.date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Styling
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    gold_font = Font(color="D4AF37", bold=True)
    
    # Header
    ws['A1'] = "Transaction List"
    ws['A1'].font = gold_font
    ws['A2'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Column headers
    headers = ['Date', 'Description', 'Category', 'Amount (RM)', 'Type', 'Linked Account']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = header
        cell.fill = gold_fill
        cell.font = white_font
    
    # Data rows
    row = 5
    for tx in transactions:
        ws[f'A{row}'] = tx.date.strftime('%Y-%m-%d')
        ws[f'B{row}'] = tx.description
        ws[f'C{row}'] = tx.category or 'Uncategorized'
        ws[f'D{row}'] = tx.amount
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'E{row}'] = 'Income' if tx.amount > 0 else 'Expense'
        
        # Get linked account name
        if tx.account_id:
            account = BalanceItem.query.get(tx.account_id)
            ws[f'F{row}'] = account.name if account else 'N/A'
        else:
            ws[f'F{row}'] = 'Unlinked'
        row += 1
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'transactions_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/balance_sheet/export')
@login_required
def balance_sheet_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import send_file
    import datetime
    
    # Get balance sheet data (reuse logic from balance_sheet route)
    transactions = Transaction.query.all()
    total_cash_auto = sum(t.amount for t in transactions if t.account_id is None)
    
    manual_items = BalanceItem.query.all()
    account_sums = db.session.query(
        Transaction.account_id, func.sum(Transaction.amount)
    ).filter(Transaction.account_id.isnot(None)).group_by(Transaction.account_id).all()
    account_sum_map = {acc_id: total for acc_id, total in account_sums}
    
    # Build assets and liabilities
    assets = {'current': [], 'non_current': []}
    liabilities = {'current': [], 'non_current': []}
    
    for item in manual_items:
        current_value = item.value
        transaction_sum = account_sum_map.get(item.id, 0.0)
        final_value = current_value + transaction_sum
        
        entry = {'name': item.name, 'value': final_value, 'type': item.asset_type or 'Other'}
        
        if item.classification == 'Current Asset':
            assets['current'].append(entry)
        elif item.classification == 'Non-Current Asset':
            assets['non_current'].append(entry)
        elif item.classification == 'Current Liability':
            liabilities['current'].append(entry)
        elif item.classification == 'Non-Current Liability':
            liabilities['non_current'].append(entry)
    
    # Calculate totals
    total_current_assets = sum(item['value'] for item in assets['current']) + total_cash_auto
    total_non_current_assets = sum(item['value'] for item in assets['non_current'])
    total_assets = total_current_assets + total_non_current_assets
    
    total_current_liabilities = sum(item['value'] for item in liabilities['current'])
    total_non_current_liabilities = sum(item['value'] for item in liabilities['non_current'])
    total_liabilities = total_current_liabilities + total_non_current_liabilities
    
    net_worth = total_assets - total_liabilities
    
    # Create workbook
    wb = Workbook()
    
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    gold_font = Font(color="D4AF37", bold=True)
    
    # Summary sheet
    ws1 = wb.active
    ws1.title = "Summary"
    
    ws1['A1'] = "Balance Sheet Summary"
    ws1['A1'].font = Font(size=16, bold=True, color="D4AF37")
    ws1['A2'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    ws1['A4'] = "Metric"
    ws1['B4'] = "Amount (RM)"
    ws1['A4'].fill = gold_fill
    ws1['B4'].fill = gold_fill
    ws1['A4'].font = white_font
    ws1['B4'].font = white_font
    
    summary_data = [
        ("Total Assets", total_assets),
        ("  - Current Assets", total_current_assets),
        ("  - Non-Current Assets", total_non_current_assets),
        ("", ""),
        ("Total Liabilities", total_liabilities),
        ("  - Current Liabilities", total_current_liabilities),
        ("  - Non-Current Liabilities", total_non_current_liabilities),
        ("", ""),
        ("Net Worth", net_worth)
    ]
    
    for idx, (metric, value) in enumerate(summary_data, start=5):
        ws1[f'A{idx}'] = metric
        if value != "":
            ws1[f'B{idx}'] = value
            ws1[f'B{idx}'].number_format = '#,##0.00'
            if metric == "Net Worth":
                ws1[f'B{idx}'].font = Font(bold=True, color="D4AF37")
    
    # Assets sheet
    ws2 = wb.create_sheet("Assets")
    ws2['A1'] = "Assets Breakdown"
    ws2['A1'].font = gold_font
    
    ws2['A3'] = "Classification"
    ws2['B3'] = "Item"
    ws2['C3'] = "Type"
    ws2['D3'] = "Amount (RM)"
    for cell in [ws2['A3'], ws2['B3'], ws2['C3'], ws2['D3']]:
        cell.fill = gold_fill
        cell.font = white_font
    
    row = 4
    for item in assets['current']:
        ws2[f'A{row}'] = "Current"
        ws2[f'B{row}'] = item['name']
        ws2[f'C{row}'] = item['type']
        ws2[f'D{row}'] = item['value']
        ws2[f'D{row}'].number_format = '#,##0.00'
        row += 1
    
    if total_cash_auto > 0:
        ws2[f'A{row}'] = "Current"
        ws2[f'B{row}'] = "Unallocated Cash"
        ws2[f'C{row}'] = "Cash"
        ws2[f'D{row}'] = total_cash_auto
        ws2[f'D{row}'].number_format = '#,##0.00'
        row += 1
    
    for item in assets['non_current']:
        ws2[f'A{row}'] = "Non-Current"
        ws2[f'B{row}'] = item['name']
        ws2[f'C{row}'] = item['type']
        ws2[f'D{row}'] = item['value']
        ws2[f'D{row}'].number_format = '#,##0.00'
        row += 1
    
    # Liabilities sheet
    ws3 = wb.create_sheet("Liabilities")
    ws3['A1'] = "Liabilities Breakdown"
    ws3['A1'].font = gold_font
    
    ws3['A3'] = "Classification"
    ws3['B3'] = "Item"
    ws3['C3'] = "Type"
    ws3['D3'] = "Amount (RM)"
    for cell in [ws3['A3'], ws3['B3'], ws3['C3'], ws3['D3']]:
        cell.fill = gold_fill
        cell.font = white_font
    
    row = 4
    for item in liabilities['current']:
        ws3[f'A{row}'] = "Current"
        ws3[f'B{row}'] = item['name']
        ws3[f'C{row}'] = item['type']
        ws3[f'D{row}'] = item['value']
        ws3[f'D{row}'].number_format = '#,##0.00'
        row += 1
    
    for item in liabilities['non_current']:
        ws3[f'A{row}'] = "Non-Current"
        ws3[f'B{row}'] = item['name']
        ws3[f'C{row}'] = item['type']
        ws3[f'D{row}'] = item['value']
        ws3[f'D{row}'].number_format = '#,##0.00'
        row += 1
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'balance_sheet_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/portfolio/export')
@login_required
def portfolio_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import send_file
    import datetime
    
    investments = Investment.query.order_by(Investment.date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Investment Portfolio"
    
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    gold_font = Font(color="D4AF37", bold=True)
    
    ws['A1'] = "Investment Portfolio"
    ws['A1'].font = gold_font
    ws['A2'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Headers
    headers = ['Date', 'Symbol', 'Type', 'Quantity', 'Price', 'Fees', 'Currency', 'Exchange Rate', 'Total (MYR)', 'Remark']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = header
        cell.fill = gold_fill
        cell.font = white_font
    
    # Data
    row = 5
    for inv in investments:
        ws[f'A{row}'] = inv.date.strftime('%Y-%m-%d')
        ws[f'B{row}'] = inv.symbol
        ws[f'C{row}'] = inv.type
        ws[f'D{row}'] = inv.quantity
        ws[f'E{row}'] = inv.price
        ws[f'E{row}'].number_format = '#,##0.00'
        ws[f'F{row}'] = inv.fees
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'G{row}'] = inv.currency
        ws[f'H{row}'] = inv.exchange_rate
        ws[f'H{row}'].number_format = '#,##0.0000'
        
        # Calculate total in MYR
        if inv.type == 'Buy':
            total_myr = ((inv.price * inv.quantity) + inv.fees) * inv.exchange_rate
        elif inv.type == 'Sell':
            total_myr = ((inv.price * inv.quantity) - inv.fees) * inv.exchange_rate
        else:
            total_myr = 0
        
        ws[f'I{row}'] = total_myr
        ws[f'I{row}'].number_format = '#,##0.00'
        ws[f'J{row}'] = inv.remark or ''
        row += 1
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'portfolio_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


# ===== EXCEL IMPORT FUNCTIONALITY =====

@app.route('/download_transaction_template')
@login_required
def download_transaction_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import send_file
    import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Template"
    
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    # Instructions
    ws['A1'] = "Transaction Import Template"
    ws['A1'].font = Font(size=14, bold=True, color="D4AF37")
    ws['A2'] = "Instructions: Fill in your transaction data below. Date format: YYYY-MM-DD. Amount: negative for expenses, positive for income."
    
    # Headers
    headers = ['Date', 'Description', 'Category', 'Amount']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = header
        cell.fill = gold_fill
        cell.font = white_font
    
    # Example rows
    ws['A5'] = '2024-01-15'
    ws['B5'] = 'Monthly Salary'
    ws['C5'] = 'Income'
    ws['D5'] = 5000.00
    
    ws['A6'] = '2024-01-16'
    ws['B6'] = 'Grocery Shopping'
    ws['C6'] = 'Food'
    ws['D6'] = -150.50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='transaction_template.xlsx'
    )


@app.route('/import_transactions', methods=['POST'])
@login_required
def import_transactions():
    from openpyxl import load_workbook
    from flask import request, flash, redirect, url_for
    import datetime as dt
    
    if 'file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('index'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('index'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'danger')
        return redirect(url_for('index'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        # Find header row (look for "Date" column)
        header_row = None
        for row_idx in range(1, 10):
            if ws.cell(row=row_idx, column=1).value in ['Date', 'date', 'DATE']:
                header_row = row_idx
                break
        
        if header_row is None:
            flash('Could not find header row with "Date" column', 'danger')
            return redirect(url_for('index'))
        
        # Parse rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                date_val = ws.cell(row=row_idx, column=1).value
                description = ws.cell(row=row_idx, column=2).value
                category = ws.cell(row=row_idx, column=3).value
                amount = ws.cell(row=row_idx, column=4).value
                
                # Skip empty rows
                if not date_val or not description or amount is None:
                    continue
                
                # Parse date
                if isinstance(date_val, dt.datetime):
                    date_obj = date_val.date()
                elif isinstance(date_val, dt.date):
                    date_obj = date_val
                else:
                    date_obj = dt.datetime.strptime(str(date_val), '%Y-%m-%d').date()
                
                # Convert amount to float
                amount_float = float(amount)
                
                # Duplicate detection
                existing = Transaction.query.filter_by(
                    date=date_obj,
                    description=description,
                    amount=amount_float
                ).first()
                
                if existing:
                    skip_count += 1
                    continue
                
                # Create transaction
                transaction = Transaction(
                    date=date_obj,
                    description=description,
                    category=category or 'Uncategorized',
                    amount=amount_float
                )
                db.session.add(transaction)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"Row {row_idx}: {str(e)}")
        
        db.session.commit()
        
        # Flash results
        if success_count > 0:
            flash(f'Successfully imported {success_count} transactions', 'success')
        if skip_count > 0:
            flash(f'Skipped {skip_count} duplicate transactions', 'warning')
        if error_count > 0:
            flash(f'{error_count} errors occurred. See details below.', 'danger')
            for error in errors[:5]:  # Show first 5 errors
                flash(error, 'danger')
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'danger')
    
    return redirect(url_for('index'))


@app.route('/download_balance_sheet_template')
@login_required
def download_balance_sheet_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import send_file
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet Template"
    
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    ws['A1'] = "Balance Sheet Import Template"
    ws['A1'].font = Font(size=14, bold=True, color="D4AF37")
    ws['A2'] = "Instructions: Classification = Current Asset/Non-Current Asset/Current Liability/Non-Current Liability"
    
    headers = ['Classification', 'Name', 'Value', 'Type', 'Liquidity Tier', 'Obligation Type']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = header
        cell.fill = gold_fill
        cell.font = white_font
    
    # Examples
    ws['A5'] = 'Current Asset'
    ws['B5'] = 'MBB Savings'
    ws['C5'] = 10000.00
    ws['D5'] = 'Cash'
    ws['E5'] = 'High'
    ws['F5'] = ''
    
    ws['A6'] = 'Current Liability'
    ws['B6'] = 'Credit Card MBB'
    ws['C6'] = 2000.00
    ws['D6'] = 'Credit Card'
    ws['E6'] = ''
    ws['F6'] = 'Standard'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='balance_sheet_template.xlsx'
    )


@app.route('/import_balance_sheet', methods=['POST'])
@login_required
def import_balance_sheet():
    from openpyxl import load_workbook
    from flask import request, flash, redirect, url_for
    
    if 'file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('balance_sheet'))
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'danger')
        return redirect(url_for('balance_sheet'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        skip_count = 0
        errors = []
        
        # Find header row
        header_row = None
        for row_idx in range(1, 10):
            if ws.cell(row=row_idx, column=1).value in ['Classification', 'classification']:
                header_row = row_idx
                break
        
        if header_row is None:
            flash('Could not find header row', 'danger')
            return redirect(url_for('balance_sheet'))
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                classification = ws.cell(row=row_idx, column=1).value
                name = ws.cell(row=row_idx, column=2).value
                value = ws.cell(row=row_idx, column=3).value
                asset_type = ws.cell(row=row_idx, column=4).value
                liquidity_tier = ws.cell(row=row_idx, column=5).value
                obligation_type = ws.cell(row=row_idx, column=6).value
                
                if not classification or not name or value is None:
                    continue
                
                # Duplicate detection
                existing = BalanceItem.query.filter_by(
                    classification=classification,
                    name=name
                ).first()
                
                if existing:
                    skip_count += 1
                    continue
                
                item = BalanceItem(
                    classification=classification,
                    name=name,
                    value=float(value),
                    asset_type=asset_type if asset_type else None,
                    liquidity_tier=liquidity_tier if liquidity_tier else None,
                    obligation_type=obligation_type if obligation_type else 'Standard'
                )
                db.session.add(item)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        db.session.commit()
        
        if success_count > 0:
            flash(f'Successfully imported {success_count} items', 'success')
        if skip_count > 0:
            flash(f'Skipped {skip_count} duplicates', 'warning')
        if errors:
            flash(f'{len(errors)} errors occurred', 'danger')
            for error in errors[:3]:
                flash(error, 'danger')
                
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('balance_sheet'))


@app.route('/download_portfolio_template')
@login_required
def download_portfolio_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import send_file
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Template"
    
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    ws['A1'] = "Portfolio Import Template"
    ws['A1'].font = Font(size=14, bold=True, color="D4AF37")
    ws['A2'] = "Instructions: Type = Buy/Sell/Dividend/Bonus/Split. Currency = MYR/USD. Date format: YYYY-MM-DD"
    
    headers = ['Date', 'Symbol', 'Type', 'Quantity', 'Price', 'Fees', 'Currency', 'Exchange Rate', 'Remark']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = header
        cell.fill = gold_fill
        cell.font = white_font
    
    # Examples
    ws['A5'] = '2024-01-15'
    ws['B5'] = 'AAPL'
    ws['C5'] = 'Buy'
    ws['D5'] = 10
    ws['E5'] = 150.00
    ws['F5'] = 5.00
    ws['G5'] = 'USD'
    ws['H5'] = 4.50
    ws['I5'] = 'Initial purchase'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='portfolio_template.xlsx'
    )


@app.route('/import_portfolio', methods=['POST'])
@login_required
def import_portfolio():
    from openpyxl import load_workbook
    from flask import request, flash, redirect, url_for
    import datetime as dt
    
    if 'file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('portfolio'))
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'danger')
        return redirect(url_for('portfolio'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        skip_count = 0
        errors = []
        
        # Find header
        header_row = None
        for row_idx in range(1, 10):
            if ws.cell(row=row_idx, column=1).value in ['Date', 'date']:
                header_row = row_idx
                break
        
        if header_row is None:
            flash('Could not find header row', 'danger')
            return redirect(url_for('portfolio'))
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                date_val = ws.cell(row=row_idx, column=1).value
                symbol = ws.cell(row=row_idx, column=2).value
                inv_type = ws.cell(row=row_idx, column=3).value
                quantity = ws.cell(row=row_idx, column=4).value
                price = ws.cell(row=row_idx, column=5).value
                fees = ws.cell(row=row_idx, column=6).value
                currency = ws.cell(row=row_idx, column=7).value
                exchange_rate = ws.cell(row=row_idx, column=8).value
                remark = ws.cell(row=row_idx, column=9).value
                
                if not date_val or not symbol or not inv_type:
                    continue
                
                # Parse date
                if isinstance(date_val, dt.datetime):
                    date_obj = date_val.date()
                elif isinstance(date_val, dt.date):
                    date_obj = date_val
                else:
                    date_obj = dt.datetime.strptime(str(date_val), '%Y-%m-%d').date()
                
                # Duplicate detection
                existing = Investment.query.filter_by(
                    date=date_obj,
                    symbol=symbol,
                    type=inv_type,
                    quantity=float(quantity or 0),
                    price=float(price or 0)
                ).first()
                
                if existing:
                    skip_count += 1
                    continue
                
                investment = Investment(
                    date=date_obj,
                    symbol=symbol,
                    type=inv_type,
                    quantity=float(quantity or 0),
                    price=float(price or 0),
                    fees=float(fees or 0),
                    currency=currency or 'MYR',
                    exchange_rate=float(exchange_rate or 1.0),
                    remark=remark
                )
                db.session.add(investment)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        db.session.commit()
        
        if success_count > 0:
            flash(f'Successfully imported {success_count} investments', 'success')
        if skip_count > 0:
            flash(f'Skipped {skip_count} duplicates', 'warning')
        if errors:
            flash(f'{len(errors)} errors occurred', 'danger')
            for error in errors[:3]:
                flash(error, 'danger')
                
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('portfolio'))


@app.route('/balance_sheet/delete/<int:id>', methods=['POST'])
@login_required
def delete_balance_item(id):
    item = BalanceItem.query.get_or_404(id)
    
    # Unlink any transactions associated with this item
    linked_transactions = Transaction.query.filter_by(account_id=id).all()
    for t in linked_transactions:
        t.account_id = None
        
    db.session.delete(item)
    db.session.commit()
    return redirect(url_for('balance_sheet'))

@app.route('/balance_sheet/edit/<int:id>', methods=['POST'])
@login_required
def edit_balance_item(id):
    item = BalanceItem.query.get_or_404(id)
    item.name = request.form['name']
    
    # Handle Liquidity Tier
    tier = request.form.get('liquidity_tier')
    if tier in ['High', 'Medium', 'Low']:
        item.liquidity_tier = tier
    else:
        item.liquidity_tier = None
        
    try:
        val_str = request.form['value']
        if not val_str:
            new_value = 0.0
        else:
            new_value = float(val_str)
        
        # Calculate current transaction sum for this account
        transaction_sum = db.session.query(func.sum(Transaction.amount)).filter_by(account_id=id).scalar() or 0.0
        
        # The user sees: stored_value + transaction_sum
        # The user inputs new_value as the FINAL displayed value
        # So we need to store: new_value - transaction_sum
        base_value = new_value - transaction_sum
        
        # Contra Logic uses the DISPLAYED values (includes transactions)
        if 'enable_contra' in request.form and request.form.get('contra_item_id'):
            contra_id = request.form.get('contra_item_id')
            
            # Calculate diff based on submitted OLD value (BF) if present, else displayed value
            if request.form.get('old_value'):
                old_displayed_value = float(request.form['old_value'])
            else:
                # Old displayed value = old base + transactions
                old_displayed_value = item.value + transaction_sum
            
            diff = new_value - old_displayed_value
            
            # Apply reverse adjustment to contra account
            contra_item = BalanceItem.query.get(contra_id)
            if contra_item:
                # Helper for asset check
                def check_asset(classification):
                    return 'Asset' in classification or 'Receivable' in classification

                is_asset_item = check_asset(item.classification)
                is_asset_contra = check_asset(contra_item.classification)

                app.logger.info(f"DEBUG: Item={item.name}, Base(DB)={item.value}, TxnSum={transaction_sum}, Old(Displayed)={old_displayed_value}, New(Displayed)={new_value}, Diff={diff}")
                app.logger.info(f"DEBUG: Contra={contra_item.name}, Class={contra_item.classification}, Val={contra_item.value}")
                app.logger.info(f"DEBUG: IsAsset(Item)={is_asset_item}, IsAsset(Contra)={is_asset_contra}")

                if is_asset_item == is_asset_contra:
                    app.logger.info("DEBUG: Condition TRUE (Same Type). Subtracting Diff.")
                    contra_item.value -= diff
                else:
                    app.logger.info("DEBUG: Condition FALSE (Diff Type). Adding Diff.")
                    contra_item.value += diff
                app.logger.info(f"DEBUG: Final Contra Val={contra_item.value}")

        # Store the BASE value (transactions will be added automatically on display)
        old_base_value = item.value
        item.value = base_value
        app.logger.info(f"DEBUG: Storing base_value={base_value} for {item.name} (displayed will be {base_value + transaction_sum})")
        
        # Record this change in history
        contra_name = None
        contra_id_for_history = None
        if 'enable_contra' in request.form and request.form.get('contra_item_id'):
            contra_id_for_history = int(request.form.get('contra_item_id'))
            contra_item_for_name = BalanceItem.query.get(contra_id_for_history)
            if contra_item_for_name:
                contra_name = contra_item_for_name.name
        
        history_entry = BalanceHistory(
            item_id=id,
            old_value=old_base_value + transaction_sum,  # Store displayed values
            new_value=new_value,  # This is what user sees
            adjustment=new_value - (old_base_value + transaction_sum),
            contra_account_id=contra_id_for_history,
            contra_account_name=contra_name,
            description=f"Manual edit" + (f" → Transfer to {contra_name}" if contra_name else "")
        )
        db.session.add(history_entry)
    except ValueError:
        pass
    db.session.commit()
    return redirect(url_for('balance_sheet'))

@app.route('/balance_sheet/transactions/<int:id>')
@login_required
def get_balance_item_transactions(id):
    transactions = Transaction.query.filter_by(account_id=id).order_by(Transaction.date.desc()).all()
    return jsonify([{
        'date': t.date.strftime('%Y-%m-%d'),
        'description': t.description,
        'amount': t.amount,
        'category': t.category
    } for t in transactions])

@app.route('/balance_sheet/history/<int:id>')
@login_required
def get_balance_item_history(id):
    history = BalanceHistory.query.filter_by(item_id=id).order_by(BalanceHistory.date.desc()).limit(20).all()
    return jsonify([{
        'date': h.date.strftime('%Y-%m-%d %H:%M'),
        'old_value': h.old_value,
        'new_value': h.new_value,
        'adjustment': h.adjustment,
        'contra_account': h.contra_account_name,
        'description': h.description
    } for h in history])

@app.route('/balance_sheet/add_mortgage', methods=['POST'])
@login_required
def add_mortgage():
    name = request.form['name']
    start_date_str = request.form['start_date']
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = datetime.utcnow().date()
        
    original_principal = float(request.form['original_principal'])
    term_years = int(request.form['term_years'])
    initial_rate = float(request.form['initial_rate'])
    
    has_mrta = 'has_mrta' in request.form
    mrta_original_amount = float(request.form.get('mrta_original_amount', 0)) if has_mrta else 0
    mrta_rate = float(request.form.get('mrta_rate', 0)) if has_mrta else 0
    
    # Create Mortgage
    new_mortgage = Mortgage(
        name=name, start_date=start_date, original_principal=original_principal,
        term_years=term_years, has_mrta=has_mrta,
        mrta_original_amount=mrta_original_amount, mrta_rate=mrta_rate
    )
    db.session.add(new_mortgage)
    db.session.flush() # Get ID
    
    # Create Initial Rate Event
    # We record the initial rate as an event on the start date
    rate_event = MortgageEvent(
        mortgage_id=new_mortgage.id,
        date=start_date,
        type='RATE_CHANGE',
        value=initial_rate,
        balance_after=original_principal # Initial balance matches principal (no payment yet)
    )
    db.session.add(rate_event)
    db.session.commit()
    
    return redirect(url_for('balance_sheet'))

@app.route('/portfolio/overview')
@login_required
def portfolio_overview():
    filter_type = request.args.get('filter', 'overall')
    view_currency = request.args.get('currency', 'MYR')  # MYR, USD, or Original
    
    investments = Investment.query.order_by(Investment.date).all()
    avg_forex_rate = get_average_forex_rate()
    stock_prices = {sp.symbol: sp.price for sp in StockPrice.query.all()}
    
    # Get current USD/MYR rate for distinct P&L calculation (Market Value)
    usd_myr_price = StockPrice.query.get('USDMYR')
    current_forex_rate = usd_myr_price.price if usd_myr_price else avg_forex_rate

    # Track detailed position information
    positions = {}  # symbol -> position data
    
    for inv in investments:
        if inv.symbol not in positions:
            positions[inv.symbol] = {
                'market': inv.market,
                'currency': inv.currency,
                'quantity': 0,
                'cost_basis': 0,
                'total_dividends': 0,
                'realized_pnl': 0,
                'first_buy_date': None,
                'last_sell_date': None,
                'total_bought': 0,
                'total_sold': 0,
                'avg_buy_price': 0,
                'avg_sell_price': 0,
                'buy_cost': 0,
                'sell_proceeds': 0
            }
        
        pos = positions[inv.symbol]
        
        if inv.type == 'Buy':
            if pos['first_buy_date'] is None:
                pos['first_buy_date'] = inv.date
            pos['quantity'] += inv.quantity
            pos['total_bought'] += inv.quantity
            buy_cost = (inv.price * inv.quantity) + inv.fees
            pos['cost_basis'] += buy_cost
            pos['buy_cost'] += buy_cost
            
        elif inv.type == 'Sell':
            if pos['quantity'] > 0:
                # Calculate average cost per share
                avg_cost_per_share = pos['cost_basis'] / pos['quantity']
                cost_removed = avg_cost_per_share * inv.quantity
                sale_proceeds = (inv.price * inv.quantity) - inv.fees
                realized_pnl = sale_proceeds - cost_removed
                
                pos['realized_pnl'] += realized_pnl
                pos['cost_basis'] -= cost_removed
                pos['quantity'] -= inv.quantity
                pos['total_sold'] += inv.quantity
                pos['sell_proceeds'] += sale_proceeds
                pos['last_sell_date'] = inv.date
                
        elif inv.type in ['Bonus', 'Split']:
            pos['quantity'] += inv.quantity
            pos['total_bought'] += inv.quantity
            
        elif inv.type == 'Dividend':
            pos['total_dividends'] += inv.price
    
    # Calculate average prices
    for symbol, pos in positions.items():
        if pos['total_bought'] > 0:
            pos['avg_buy_price'] = pos['buy_cost'] / pos['total_bought']
        if pos['total_sold'] > 0:
            pos['avg_sell_price'] = pos['sell_proceeds'] / pos['total_sold']
    
    # Separate current holdings and sold positions
    current_holdings = {k: v for k, v in positions.items() if v['quantity'] > 0}
    sold_positions = {k: v for k, v in positions.items() if v['quantity'] == 0 and v['total_sold'] > 0}
    
    # Calculate metrics based on filter
    total_invested = 0
    total_market_value = 0
    total_dividends = 0
    total_realized_pnl = 0
    market_allocation = {}
    currency_allocation = {'MYR': 0, 'USD': 0}
    detailed_items = []
    
    # Select positions to process
    if filter_type == 'holdings':
        items_to_process = current_holdings
    elif filter_type == 'sold':
        items_to_process = sold_positions
    else:  # overall
        items_to_process = positions
    
    for symbol, pos in items_to_process.items():
        current_price = stock_prices.get(symbol, 0.0)
        market_value = pos['quantity'] * current_price
        
        # Convert based on view_currency
        if view_currency == 'MYR':
            # Convert everything to MYR
            if pos['currency'] == 'USD':
                cost_display = pos['cost_basis'] * avg_forex_rate
                value_display = market_value * current_forex_rate
                div_display = pos['total_dividends'] * avg_forex_rate
                realized_pnl_display = pos['realized_pnl'] * avg_forex_rate
                avg_buy_display = pos['avg_buy_price'] * avg_forex_rate
                avg_sell_display = pos['avg_sell_price'] * avg_forex_rate
            else:
                cost_display = pos['cost_basis']
                value_display = market_value
                div_display = pos['total_dividends']
                realized_pnl_display = pos['realized_pnl']
                avg_buy_display = pos['avg_buy_price']
                avg_sell_display = pos['avg_sell_price']
            display_currency = 'MYR'
            
        elif view_currency == 'USD':
            # Convert everything to USD
            if pos['currency'] == 'MYR':
                rate = 1.0 / avg_forex_rate if avg_forex_rate > 0 else 1.0
                cost_display = pos['cost_basis'] * rate
                value_display = market_value * rate
                div_display = pos['total_dividends'] * rate
                realized_pnl_display = pos['realized_pnl'] * rate
                avg_buy_display = pos['avg_buy_price'] * rate
                avg_sell_display = pos['avg_sell_price'] * rate
            else:
                cost_display = pos['cost_basis']
                value_display = market_value
                div_display = pos['total_dividends']
                realized_pnl_display = pos['realized_pnl']
                avg_buy_display = pos['avg_buy_price']
                avg_sell_display = pos['avg_sell_price']
            display_currency = 'USD'
            
        else:  # Original
            # Keep in original currency
            cost_display = pos['cost_basis']
            value_display = market_value
            div_display = pos['total_dividends']
            realized_pnl_display = pos['realized_pnl']
            avg_buy_display = pos['avg_buy_price']
            avg_sell_display = pos['avg_sell_price']
            display_currency = pos['currency']
        
        # For totals, always use MYR (for allocation/summary)
        if pos['currency'] == 'USD':
            cost_myr = pos['cost_basis'] * avg_forex_rate
            value_myr = market_value * current_forex_rate
            div_myr = pos['total_dividends'] * avg_forex_rate
            realized_pnl_myr = pos['realized_pnl'] * avg_forex_rate
        else:
            cost_myr = pos['cost_basis']
            value_myr = market_value
            div_myr = pos['total_dividends']
            realized_pnl_myr = pos['realized_pnl']
        
        total_invested += cost_myr
        total_market_value += value_myr
        total_dividends += div_myr
        total_realized_pnl += realized_pnl_myr
        
        # Market/Currency allocation (only for current holdings)
        if pos['quantity'] > 0:
            if pos['market'] not in market_allocation:
                market_allocation[pos['market']] = 0
            market_allocation[pos['market']] += value_myr
            
            if pos['currency'] == 'USD':
                currency_allocation['USD'] += value_myr
            else:
                currency_allocation['MYR'] += value_myr
        
        # Calculate P&L
        unrealized_pnl = value_display - cost_display
        total_pnl_display = unrealized_pnl + realized_pnl_display
        
        # Calculate P&L percentage
        if pos['quantity'] == 0:  # Sold position
            original_cost = pos['buy_cost']
            if pos['currency'] == 'USD' and view_currency == 'MYR':
                original_cost *= avg_forex_rate
            elif pos['currency'] == 'MYR' and view_currency == 'USD':
                original_cost *= (1.0 / avg_forex_rate if avg_forex_rate > 0 else 1.0)
            
            if original_cost > 0:
                pnl_percent = ((realized_pnl_display + div_display) / original_cost) * 100
            else:
                pnl_percent = 0
        else:  # Current holding
            if cost_display > 0:
                pnl_percent = (total_pnl_display / cost_display) * 100
            else:
                pnl_percent = 0
        
        detailed_items.append({
            'symbol': symbol,
            'market': pos['market'],
            'quantity': pos['quantity'],
            'total_bought': pos['total_bought'],
            'total_sold': pos['total_sold'],
            'avg_buy_price': avg_buy_display,
            'avg_sell_price': avg_sell_display if pos['total_sold'] > 0 else 0,
            'cost': cost_display,
            'value': value_display,
            'dividends': div_display,
            'pnl_value': total_pnl_display,
            'pnl_percent': pnl_percent,
            'first_buy_date': pos['first_buy_date'],
            'last_sell_date': pos['last_sell_date'],
            'status': 'Sold' if pos['quantity'] == 0 else 'Holding',
            'currency': display_currency
        })
    
    # Sort by absolute P&L
    detailed_items.sort(key=lambda x: abs(x['pnl_value']), reverse=True)
    
    # Calculate total P&L
    total_unrealized_pnl = total_market_value - total_invested
    total_pnl = total_unrealized_pnl + total_realized_pnl
    
    # Calculate overall P&L percentage
    if total_invested + total_realized_pnl > 0:
         total_pnl_percent = (total_pnl / (total_invested + total_realized_pnl)) * 100
    else:
         total_pnl_percent = 0
    
    # Convert summary totals to display currency
    if view_currency == 'USD':
        rate = 1.0 / avg_forex_rate if avg_forex_rate > 0 else 1.0
        total_invested_display = total_invested * rate
        total_market_value_display = total_market_value * rate
        total_pnl_display = total_pnl * rate
        total_dividends_display = total_dividends * rate
        total_realized_pnl_display = total_realized_pnl * rate
        display_currency = 'USD'
    else:  # MYR or Original (summary is always in one currency)
        total_invested_display = total_invested
        total_market_value_display = total_market_value
        total_pnl_display = total_pnl
        total_dividends_display = total_dividends
        total_realized_pnl_display = total_realized_pnl
        display_currency = 'MYR'
    
    return render_template('portfolio_overview.html',
                         total_invested=total_invested_display,
                         total_market_value=total_market_value_display,
                         total_pnl=total_pnl_display,
                         total_pnl_percent=total_pnl_percent,
                         total_dividends=total_dividends_display,
                         total_realized_pnl=total_realized_pnl_display,
                         display_currency=display_currency,
                         market_allocation=market_allocation,
                         currency_allocation=currency_allocation,
                         detailed_items=detailed_items,
                         filter_type=filter_type,
                         view_currency=view_currency,
                         avg_forex_rate=avg_forex_rate,
                         current_forex_rate=current_forex_rate,
                         holdings_count=len(current_holdings),
                         sold_count=len(sold_positions))

@app.template_filter('comma')
def comma_filter(value):
    try:
        return "{:,.2f}".format(value)
    except (ValueError, TypeError):
        return value

# ADMIN: Reset All Data
@app.route('/admin/reset_all_data', methods=['POST'])
@login_required
def reset_all_data():
    try:
        app.logger.info("=== RESET ALL DATA STARTED ===")
        
        # Count records before deletion
        count_history = db.session.query(BalanceHistory).count()
        count_txn = db.session.query(Transaction).count()
        count_inv = db.session.query(Investment).count()
        count_bal = db.session.query(BalanceItem).count()
        
        app.logger.info(f"Before: History={count_history}, Txn={count_txn}, Inv={count_inv}, Bal={count_bal}")
        
        # Delete all data from all tables
        db.session.query(BalanceHistory).delete()
        db.session.query(Transaction).delete()
        db.session.query(Investment).delete()
        db.session.query(MortgageEvent).delete()
        db.session.query(Mortgage).delete()
        db.session.query(BalanceItem).delete()
        db.session.query(StockPrice).delete()
        db.session.query(ForexTransaction).delete()
        
        db.session.commit()
        
        # Count records after deletion
        count_history_after = db.session.query(BalanceHistory).count()
        count_txn_after = db.session.query(Transaction).count()
        count_inv_after = db.session.query(Investment).count()
        count_bal_after = db.session.query(BalanceItem).count()
        
        app.logger.info(f"After: History={count_history_after}, Txn={count_txn_after}, Inv={count_inv_after}, Bal={count_bal_after}")
        app.logger.info("=== RESET ALL DATA COMPLETED SUCCESSFULLY ===")
        
        flash('All data has been successfully deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"=== RESET ALL DATA FAILED: {str(e)} ===")
        flash(f'Error resetting data: {str(e)}', 'danger')
    
    return redirect(url_for('balance_sheet'))


# ===== AR EMAIL REMINDER ROUTES =====

from email_config import send_payment_reminder, test_email_config, load_email_config, save_email_config

@app.route('/ar/send_reminder/<int:id>', methods=['POST'])
@login_required
def send_ar_reminder(id):
    """Send payment reminder email to single AR item"""
    item = BalanceItem.query.get_or_404(id)
    
    # Validate it's an AR item
    if item.asset_type != 'AR':
        flash('This item is not an Accounts Receivable item.', 'danger')
        return redirect(url_for('balance_sheet'))
    
    # Check if email is configured
    if not item.contact_email:
        flash(f'No email address configured for {item.name}.', 'warning')
        return redirect(url_for('balance_sheet'))
    
    # Send email
    success, message = send_payment_reminder(
        ar_name=item.name,
        to_email=item.contact_email,
        amount=item.value,
        currency='MYR',
        description=f'Account: {item.name}'
    )
    
    if success:
        # Update reminder tracking
        item.last_reminder_sent = datetime.utcnow()
        item.reminder_count = (item.reminder_count or 0) + 1
        db.session.commit()
        flash(f'✅ Payment reminder sent to {item.contact_email}', 'success')
    else:
        flash(f'❌ Failed to send email: {message}', 'danger')
    
    return redirect(url_for('balance_sheet'))

@app.route('/ar/send_bulk_reminders', methods=['POST'])
@login_required
def send_bulk_ar_reminders():
    """Send payment reminders to all AR items with emails"""
    ar_items = BalanceItem.query.filter(
        BalanceItem.asset_type == 'AR',
        BalanceItem.contact_email.isnot(None),
        BalanceItem.contact_email != ''
    ).all()
    
    if not ar_items:
        flash('No AR items with email addresses found.', 'warning')
        return redirect(url_for('balance_sheet'))
    
    sent_count = 0
    failed_count = 0
    
    for item in ar_items:
        success, message = send_payment_reminder(
            ar_name=item.name,
            to_email=item.contact_email,
            amount=item.value,
            currency='MYR',
            description=f'Account: {item.name}'
        )
        
        if success:
            item.last_reminder_sent = datetime.utcnow()
            item.reminder_count = (item.reminder_count or 0) + 1
            sent_count += 1
        else:
            failed_count += 1
    
    db.session.commit()
    
    if sent_count > 0:
        flash(f'✅ Sent {sent_count} reminder(s) successfully.', 'success')
    if failed_count > 0:
        flash(f'❌ Failed to send {failed_count} reminder(s).', 'danger')
    
    return redirect(url_for('balance_sheet'))

@app.route('/ar/email_settings', methods=['GET', 'POST'])
@login_required
def ar_email_settings():
    """Email configuration page"""
    if request.method == 'POST':
        config = {
            'email': request.form.get('email', ''),
            'app_password': request.form.get('app_password', ''),
            'smtp_server': request.form.get('smtp_server', 'smtp.gmail.com'),
            'smtp_port': int(request.form.get('smtp_port', 587)),
            'sender_name': request.form.get('sender_name', 'Finance Tracker')
        }
        save_email_config(config)
        flash('✅ Email settings saved successfully!', 'success')
        return redirect(url_for('ar_email_settings'))
    
    config = load_email_config()
    return render_template('email_settings.html', config=config)

@app.route('/ar/test_email', methods=['POST'])
@login_required
def test_ar_email():
    """Send test email"""
    test_email = request.form.get('test_email')
    
    if not test_email:
        flash('Please provide a test email address.', 'warning')
        return redirect(url_for('ar_email_settings'))
    
    success, message = test_email_config(test_email)
    
    if success:
        flash(f'✅ {message}', 'success')
    else:
        flash(f'❌ {message}', 'danger')
    
    return redirect(url_for('ar_email_settings'))

# Special route to initialize User table (visit once after deployment)
@app.route('/init_db_secret_route_12345')
def init_db():
    try:
        db.create_all()
        return "✅ Database tables created successfully! User table is ready. You can now use registration."
    except Exception as e:
        return f"❌ Error creating tables: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
